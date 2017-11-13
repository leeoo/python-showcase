# -- coding: utf-8 --

'''
Read and write excel 2007 with openpyxl.
'''

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

# 获得workbook
wb = load_workbook(filename = r'haggle.xlsx')
# 获得worksheet
ws = wb.get_sheet_by_name("Sheet4")

wb1 = Workbook()
ewb1 = ExcelWriter(workbook = wb1)
dest_filename = r'result.xlsx'
ws1 = wb1.worksheets[0]
ws1.title = "socialrange"

for i in range(0, 36):
	li = []
	for row_num in range(0, 213824):
		c2 = ws.cell(row = row_num, column = 2).value
		c3 = ws.cell(row = row_num, column = 3).value

		# 对于每个i遍历列c，找出c列跟i相同的数据，并将对应的d列的不同数据保存到li中
		if c2 == i:
			if c3 in li:
				continue
			else:
				li.append(c3)
		else:
			continue

	ws1.cell(row=i, column=0).value=i
	ws1.cell(row=i, column = 1).value = len(li)
ewb1.save(filename = dest_filename)
